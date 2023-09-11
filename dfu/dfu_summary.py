from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import pandas as pd
import numpy as np
import util.server_all_output as server_all_output


def server_db_info():
    server_db = server_all_output.output_total()
    df = pd.DataFrame()
    df["MedicalNumber"] = server_db["MedicalNumber"].to_list()
    df["site"] = df["MedicalNumber"].apply(lambda x: str(x[0:3]))
    tr_info = {"201": "nynw", "202": "ocer", "203": "whfa", "204": "youngst", "205": "lvrpool", "292": "rsci"}
    vd_info = {"206": "memdfu", "207": "hilloh", "208": "grovoh", "209": "mentoh","210": "encinogho", "211": "lahdfu"}
    total_info = {}
    for i in tr_info.keys():
        total_info[i] = tr_info[i]
    for j in vd_info.keys():
        total_info[j] = vd_info[j]
    type_list = []
    site_name_list = []
    site_list = df["site"].to_list()
    for i in site_list:
        if i in tr_info.keys():
            type_list.append("training")
            site_name_list.append(total_info[i])
        else:
            type_list.append("validation")
            site_name_list.append(total_info[i])
    df["type"] = type_list
    df["site_name"] = site_name_list
    return df


def make_summary(cur_date):
    df = server_db_info()
    df_tra = df[df["type"]=="training"]
    tra_img_total = df_tra["MedicalNumber"].to_list()
    tra_sub_image_count = Counter(tra_img_total)
    tra_img_sum = pd.DataFrame(list(tra_sub_image_count.items()), columns=["SubjectID", "Image_Num"])

    tra_sub = tra_sub_image_count.keys()
    tra_site_ab = [x[0:3] for x in tra_sub]
    tra_site_info = {}
    for i in tra_sub:
        sub_tra = tra_img_sum[tra_img_sum["SubjectID"] == i]
        num = sub_tra["Image_Num"].iloc[0]
        ii = i[0:3]
        if ii not in tra_site_info:
            tra_site_info[ii] = num
        else:
            tra_site_info[ii] += num

    tra_site_num = Counter(tra_site_ab)
    tra_site_sum = pd.DataFrame(list(tra_site_num.items()), columns=["Site", "Sub_Num"])
    tra_site_list = tra_site_sum["Site"].to_list()

    tra_sum = []
    tra_site_name = []

    for j in tra_site_list:
        tra_sub_info = df_tra[df_tra["site"] == j]
        j_name = tra_sub_info["site_name"].iloc[0]
        sum_num = tra_site_info[j]
        tra_sum.append(sum_num)
        tra_site_name.append(j_name)

    tra_site_sum["Img_Sum"] = tra_sum
    tra_site_sum["Site_Name"] = tra_site_name

    tra_total ={}
    tra_total["Site"] = "total"
    tra_total["Sub_Num"] = len(tra_sub)
    tra_total["Img_Sum"] = len(tra_img_total)

    tra_total["Site_Name"]=np.nan
    df_tra_total = pd.DataFrame(tra_total,index=[0])
    df_tra_list = [tra_site_sum,df_tra_total]
    df_tra_final = pd.concat(df_tra_list)

    # db_info = db.download_table("DFU_Master_ImageCollections")

    df_val = df[df["type"] == "validation"]
    val_img_total = df_val["MedicalNumber"].to_list()
    val_sub_image_count = Counter(val_img_total)
    val_img_sum = pd.DataFrame(list(val_sub_image_count.items()), columns=["SubjectID", "Image_Num"])


    val_sub = val_sub_image_count.keys()
    val_site_ab = [x[0:3] for x in val_sub]
    val_site_info = {}
    for i in val_sub:
        sub_val = val_img_sum[val_img_sum["SubjectID"] == i]
        num = sub_val["Image_Num"].iloc[0]
        ii = i[0:3]
        if ii not in val_site_info:
            val_site_info[ii] = num
        else:
            val_site_info[ii] += num

    val_site_num = Counter(val_site_ab)
    val_site_sum = pd.DataFrame(list(val_site_num.items()), columns=["Site", "Sub_Num"])
    val_site_list = val_site_sum["Site"].to_list()

    val_sum = []
    val_site_name = []

    for j in val_site_list:
        val_sub_info = df_val[df_val["site"] == j]
        j_name = val_sub_info["site_name"].iloc[0]
        sum_num = val_site_info[j]
        val_sum.append(sum_num)
        val_site_name.append(j_name)

    val_site_sum["Img_Sum"] = val_sum
    val_site_sum["Site_Name"] = val_site_name

    val_total = {}
    val_total["Site"] = "total"
    val_total["Sub_Num"] = len(val_sub)
    val_total["Img_Sum"] = len(val_img_total)
    val_total["Site_Name"] = np.nan

    df_val_total = pd.DataFrame(val_total,index=[0])
    df_val_list = [val_site_sum, df_val_total]
    df_val_final = pd.concat(df_val_list)

    file = "WAUSI_Summary_latest.xlsx"
    path = "../Documents/DFU_Summary/" + file



    writer = pd.ExcelWriter(path, engine='xlsxwriter')
    df_tra_final.to_excel(writer, sheet_name='training_site_summary', index=False)
    tra_img_sum.to_excel(writer, sheet_name='training_subject_summary', index=False)
    df_val_final.to_excel(writer, sheet_name='validation_site_summary', index=False)
    val_img_sum.to_excel(writer, sheet_name='validation_subject_summary', index=False)
    writer.close()
    workbook =load_workbook(path)

    # 对每个 sheet 设置单元格居中
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    # 保存更改后的 Excel 文件
    workbook.save(path)



    return path


def match_cal(match_path):
    df = pd.read_excel(match_path)
    sub_list = df["SubjectID"].to_list()
    vis_list = df["VisitTime"].to_list()
    time_list = df["Castor_Date"].to_list()
    draw_df = df[df["phase"].notna()]
    draw_df = draw_df[draw_df["phase"] != "archived"]
    draw_sub = draw_df["SubjectID"].to_list()
    draw = {}
    for i in draw_sub:
        if i not in draw:
            draw[i] = 1
        else:
            draw[i] += 1
    archive_df = df[df["phase"] == "archived"]
    archive_sub = archive_df["SubjectID"].to_list()
    archive = {}
    for i in archive_sub:
        if i not in archive:
            archive[i] = 1
        else:
            archive[i] += 1
    info = {}
    num = {}
    index = 0
    for i in sub_list:
        if i not in num:
            num[i] = 1
        else:
            num[i] += 1
        vis = vis_list[index]
        time = time_list[index]
        if i not in info:
            info[i] = {}
            if vis not in info[i]:
                info[i][vis] = time
            else:
                info[i][vis] = info[i][vis]
        else:
            if vis not in info[i]:
                info[i][vis] = time
            else:
                info[i][vis] = info[i][vis]
        index += 1
    subject = info.keys()
    column_name = ["Status","Match","not_Match","Drawn","Archive","not_Drawn"]
    sv = []
    for l in range(1, 13):
        p = "SV_" + str(l) + "_Date"
        column_name.append(p)
        sv.append(p)
    column_name.append("Visit_Total")
    df_info = pd.DataFrame(columns=column_name)
    # df_info["SubjectID"] = subject
    for i in subject:
        vis_info = info[i]
        vis_total = len(vis_info)
        for j in vis_info.keys():
            time = vis_info[j]
            df_info.loc[i, j] = time
        df_info.loc[i, "Visit_Total"] = vis_total
        try:
            df_info.loc[i, "Match"] = num[i]
        except:
            df_info.loc[i, "Match"] = 0
        try:
            df_info.loc[i, "Drawn"] = draw[i]
        except:
            df_info.loc[i, "Drawn"] = 0
        try:
            df_info.loc[i, "Archive"] = archive[i]
        except:
            df_info.loc[i, "Archive"] = 0

    df_info["not_Drawn"] = df_info["Match"]-df_info["Drawn"]-df_info["Archive"]
    return df_info,info,sv


def date_match_check(cur_date):
    df_tra = match_cal("../Documents/dfu_latest_match_training.xlsx")
    df_val= match_cal("../Documents/dfu_latest_match_validation.xlsx")
    sv = df_tra[2]

    df_tra_cal = df_tra[0]
    tra_info = df_tra[1]
    df_val_cal = df_val[0]
    val_info = df_val[1]

    column = df_tra_cal.columns.tolist()

    df_tra_cas = pd.read_excel("../Documents/Castor_training.xlsx")
    df_val_cas = pd.read_excel("../Documents/Castor_validation.xlsx")

    tra_sub = df_tra_cas["SubjectID"].to_list()
    for i in tra_sub:
        tra_cas_sub = df_tra_cas[df_tra_cas["SubjectID"]==i]
        sta = tra_cas_sub["status"].iloc[0]
        df_tra_cal.loc[i, "Status"] = sta
        for p in sv:
            cas = tra_cas_sub[p].iloc[0]
            if type(cas) != float:
                try:
                    match = tra_info[i][p]
                    df_tra_cal.loc[i, p] = match
                except:
                    df_tra_cal.loc[i,p] = "not match"
            else:
                df_tra_cal.loc[i, p] = np.nan

    df_tra_cal.reset_index(level=0, inplace=True)
    sub = ["SubjectID"]
    columns = sub + column
    df_tra_cal.columns = columns

    val_sub = df_val_cas["SubjectID"].to_list()
    for i in val_sub:
        val_cas_sub = df_val_cas[df_val_cas["SubjectID"] == i]
        sta =val_cas_sub["status"].iloc[0]
        df_val_cal.loc[i, "Status"] = sta
        for p in sv:
            cas = val_cas_sub[p].iloc[0]
            if type(cas) != float:
                try:
                    match = val_info[i][p]
                    df_val_cal.loc[i, p] = match
                except:
                    df_val_cal.loc[i, p] = "not_match"
            else:
                df_val_cal.loc[i, p] = np.nan

    df_val_cal.reset_index(level=0, inplace=True)
    sub = ["SubjectID"]
    columns = sub + column
    df_val_cal.columns = columns

    path = make_summary(cur_date)

    df_tra_site = pd.read_excel(path, sheet_name="training_site_summary")
    df_tra_sum = pd.read_excel(path,sheet_name="training_subject_summary")
    df_val_site = pd.read_excel(path,sheet_name="validation_site_summary")
    df_val_sum = pd.read_excel(path,sheet_name="validation_subject_summary")

    df_final_sum_tra = pd.merge(df_tra_sum,df_tra_cal,on="SubjectID",how="left")
    df_final_sum_val = pd.merge(df_val_sum,df_val_cal,on="SubjectID",how="left")

    df_final_sum_tra["not_Match"] = df_final_sum_tra["Image_Num"]-df_final_sum_tra["Match"]
    df_final_sum_val["not_Match"] = df_final_sum_val["Image_Num"]-df_final_sum_val["Match"]

    new_path = "../Documents/DFU_Summary/WAUSI_Summary_Calculation.xlsx"
    writer = pd.ExcelWriter(new_path, engine='xlsxwriter')
    df_tra_site.to_excel(writer, sheet_name='training_site_summary', index=False)
    df_final_sum_tra.to_excel(writer, sheet_name='training_subject_summary', index=False)
    df_val_site.to_excel(writer, sheet_name='validation_site_summary', index=False)
    df_final_sum_val.to_excel(writer, sheet_name='validation_subject_summary', index=False)
    writer.close()
    workbook = load_workbook(new_path)

    # 对每个 sheet 设置单元格居中
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    # 保存更改后的 Excel 文件
    workbook.save(new_path)

    local_name = "WAUSI_Summary_"+cur_date+".xlsx"
    local_path = "../../../Desktop/DFU_Summary/"+local_name
    workbook.save(local_path)



if __name__ =="__main__":
    date_match_check("091123")

