import os.path
import pymysql
import pandas as pd
import numpy as np
import get_sql_from_smb as sql_get
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Alignment




import download_whole_dynamodb_table

def has_element(value,element_list):
    for i in element_list:
        if i in value:
            return True

    return False

def parse_date(x):
    try:
        stamp = int(x) - 2209031999999999700 - 6*60*60*1000000000
        eval = str(pd.to_datetime(stamp).strftime('%d-%m-%Y'))
    except:
        eval = np.nan
    return eval

def server_table_output(sql_path):
    connection = pymysql.connect(
        host='127.0.0.1',
        user='root',
        password='szw970727',
        database='algorithm',
        cursorclass=pymysql.cursors.DictCursor
    )
    cursor = connection.cursor()
    path =sql_path
    with open(path, 'r') as file:
        query = file.read()

    statements = query.split(';')
    # Execute each statement
    for statement in statements:
        cursor.execute(statement)


    query1 = "select * from imagescollection left join injury on imagescollection.INJURYID=injury.INJURYID join patient on injury.PID = patient.PID"
    cursor.execute(query1)
    collection_result = cursor.fetchall()
    df = pd.DataFrame(collection_result)
    df["CaptureDate"]=df["CreateDateTime"].apply(lambda x: str(x))


    query2 = "select * from imagescollection left join images on imagescollection.IMCOLLID=images.IMCOLLID"
    cursor.execute(query2)
    image_result = cursor.fetchall()
    df2 = pd.DataFrame(image_result)
    info = sql_path.split("/")[-1]
    date = info.split("_")[-2]
    site = info.split("_")[-3]
    site_path = '/Users/ziweishi/Documents/transfer_regular_check/'+site+"/"
    if os.path.isdir(site_path)==False:
        os.mkdir(site_path)
    check_path = site_path+date+"/"
    if os.path.isdir(check_path)==False:
        os.mkdir(check_path)

    if site != "memdfu":
        element_list = ["000", "99","8888"]
    else:
        element_list = ["0000", "99"]

    df = df[~df["MedicalNumber"].apply(lambda x: has_element(x, element_list))]


    if site == "whfa":
        df["MedicalNumber"] = df["MedicalNumber"].apply(lambda x: "203-" + x)
        df["MedicalNumber"] = df["MedicalNumber"].apply(
            lambda x: x.replace("-203", "-005") if "-203" in x else x)
    if site == "lvrpool":
        df["MedicalNumber"] = df["MedicalNumber"].apply(
            lambda x: x.replace("105", "205") if "105" in x else x)
        df["MedicalNumber"] = df["MedicalNumber"].apply(
            lambda x: x.replace("205", "205-") if "205" in x else x)
    if site == "encinogho":
        df["MedicalNumber"] = df["MedicalNumber"].apply(lambda x: "210-" + x if "210" not in x else x)
    if site == "ocer":
        df["MedicalNumber"] = df["MedicalNumber"].apply(lambda x: "202-" + x if "202" not in x else x)
    if site == "youngst":
        df["MedicalNumber"] = df["MedicalNumber"].apply(
            lambda x: x.replace("204", "204-") if "204" in x else x)
        df["MedicalNumber"] = df["MedicalNumber"].apply(
            lambda x: "204-" + x if "204-" not in x else x)
    if site == "memdfu":
        df["MedicalNumber"] = df["MedicalNumber"].apply(
            lambda x: x.replace("00067469", "206-008") if "00067469" in x else x)
        df["MedicalNumber"] = df["MedicalNumber"].apply(
            lambda x: x.replace("00038523", "206-009") if "00038523" in x else x)
        df["MedicalNumber"] = df["MedicalNumber"].apply(
            lambda x: x.replace("02042426", "206-010") if "02042426" in x else x)
        df["MedicalNumber"] = df["MedicalNumber"].apply(
            lambda x: "206-" + x if "206-" not in x else x)
    if site == "lahdfu":
        df["MedicalNumber"] = df["MedicalNumber"].apply(
            lambda x: x.replace("211-01", "211-001") if "211-01" in x else x)
    if site == "rsci":
        df["MedicalNumber"] = df["MedicalNumber"].apply(
            lambda x: "292-" + x if "292-" not in x else x)

    guid_file_path = check_path+site+"_"+date+"_collection.xlsx"
    df.to_excel(guid_file_path, index=False)

    image_file_path = check_path+site+"_"+date+"_image.xlsx"
    df2.to_excel(image_file_path,index=False)

    cursor.close()
    connection.close()
    return df,df2,site,check_path


def guid_check(db,df,site,check_path):
    check_list = df["ImgCollGUID"].to_list()
    db_list = db["ImgCollGUID"].to_list()
    index=0
    issue_guid=[]
    success_guid=[]
    status=[]
    folder=[]
    subject_id=[]
    for i in check_list:
        df_sub = df[df["ImgCollGUID"] == i]
        location = df_sub["ImageCollFolderName"].iloc[0]
        subject = df_sub["MedicalNumber"].iloc[0]
        folder.append(location)
        subject_id.append(subject)
        if i in db_list:
            success_guid.append(i)
            status.append("exist")
            index+=0
        else:
            issue_guid.append(i)
            index+=1
            status.append("new")
    if index==0:
        print(site+": "+"all check!")
    else:
        print(site+": "+str(len(issue_guid))+"/"+str(len(check_list)))
    data = zip(check_list,subject_id, status,folder)
    df_guid = pd.DataFrame(data, columns=["ImgCollGUID", "SubjectID","Transfer_Status","File_Location"])
    file_name =site+"_guid_log_out.xlsx"
    status_file = os.path.join(check_path,file_name)
    df_guid.to_excel(status_file)
    return df_guid


def image_check(db,df,df2,site,check_path):
    df_guid = guid_check(db,df,site,check_path)
    guid = df_guid["ImgCollGUID"].to_list()
    status={}
    site_info=[]
    for i in guid:
        site_info.append(site)
        comment={}
        df_sub = df2[df2["ImgCollGUID"]==i]
        db_sub = db[db["ImgCollGUID"] == i]
        # check if raw in db
        if len(db_sub)==0:
            comment["Raw Issue"] = "no guid"
            comment["Pseudo Issue"] ="no guid"
            comment["Assessing Issue"] = "no guid"
            comment["Reference Issue"] = "no guid"
        else:
            # check if pseudo in db
            raw_sub = df_sub[df_sub["ImageType"] == "Raw MSI"]
            raw_list = raw_sub["ImageFileName"].to_list()
            if len(raw_list) > 0:
                try:
                    db_raw = db_sub["Raw"].iloc[0]
                    db_raw1 = db_raw.replace(" ", "")

                    db_raw1 = db_raw1.replace("Raw/Raw_", "")
                    db_raw1 = db_raw1.replace("'", "")

                    db_raw2 = db_raw1.strip("{}").split(",")

                    issue_list = []
                    for p in raw_list:
                        if p not in db_raw2:
                            issue_list.append(p)
                    if len(issue_list) > 0:
                        comment["Raw Issue"] = issue_list
                    else:
                        comment["Raw Issue"] = np.nan
                except:
                    comment["Raw Issue"] = "db no raw"

            else:
                comment["Raw Issue"] = "lc lost raw"



            # check if pseudo in db
            pseudo_sub = df_sub[df_sub["ImageType"] == "Pseudocolor"]
            pseudo_list = pseudo_sub["ImageFileName"].to_list()
            if len(pseudo_list)>0:
                try:
                    db_pseudo = db_sub["PseudoColor"].iloc[0]
                    db_pseudo1 = db_pseudo.replace("PseudoColor/", "")
                    db_pseudo1 = db_pseudo1.replace(" ", "")
                    db_pseudo1 = db_pseudo1.replace("'", "")
                    db_pseudo2 = db_pseudo1.strip("{}").split(",")
                    if len(pseudo_list) != len(db_pseudo2):
                        comment["Pseudo Issue"] = "wrong pseudo"
                    else:
                        comment["Pseudo Issue"] = np.nan
                except:
                    comment["Pseudo Issue"] = "db no pseudo"
            else:
                comment["Pseudo Issue"] = "lc lost pseudo"



            # check if assessing in db
            assessing_sub = df_sub[df_sub["ImageType"] == "CJA"]
            assessing_list = assessing_sub["ImageFileName"].to_list()
            if len(assessing_list)>0:
                try:
                    db_assessing = db_sub["Assessing"].iloc[0]
                    db_assessing1 = db_assessing.replace("Assessing/", "")
                    db_assessing1 = db_assessing1.replace(" ", "")
                    db_assessing1 = db_assessing1.replace("'", "")
                    db_assessing2 = db_assessing1.strip("{}").split(",")

                    if len(assessing_list) != len(db_assessing2):
                        comment["Assessing Issue"] = "wrong assessing"
                    else:
                        comment["Assessing Issue"] = np.nan
                except:
                    comment["Assessing Issue"] = "db no assessing"

            else:
                comment["Assessing Issue"] = "lc lost assessing"


            # check if reference in db
            reference_sub = df_sub[df_sub["ImageType"] == "Reference"]
            reference_list = reference_sub["ImageFileName"].to_list()
            if len(reference_list)>0:
                try:
                    db_reference = db_sub["Reference"].iloc[0]
                    db_reference1 = db_reference.replace("Reference/", "")
                    db_reference1 = db_reference1.replace(" ", "")
                    db_reference1 = db_reference1.replace("'", "")
                    db_reference2 = db_reference1.strip("{}").split(",")
                    for z in db_reference2:
                        if "." not in z:
                            db_reference2.remove(z)
                    if len(reference_list) != len(db_reference2):
                        comment["Reference Issue"] = "wrong reference"
                    else:
                        comment["Reference Issue"] = np.nan
                except:
                    comment["Reference Issue"] = "db no reference"
            else:
                comment["Reference Issue"] = "lc lost reference"


        status[i] = comment

    raw_issue_list=[]
    pseudo_issue_list=[]
    assessing_issue_list=[]
    reference_issue_list=[]

    for j in guid:
        issue = status[j]
        raw_issue_list.append(issue["Raw Issue"])
        pseudo_issue_list.append(issue["Pseudo Issue"])
        assessing_issue_list.append(issue["Assessing Issue"])
        reference_issue_list.append(issue["Reference Issue"])

    df_guid["Raw Issue"]=raw_issue_list
    df_guid["Pseudo Issue"] = pseudo_issue_list
    df_guid["Assessing Issue"] = assessing_issue_list
    df_guid["Reference Issue"]=reference_issue_list
    df_guid["Site"]=site_info

    file_name = site + "_image_log_out.xlsx"
    image_status_file = os.path.join(check_path, file_name)
    df_guid.to_excel(image_status_file)
    return df_guid


def refresh_sql_database():
    local_site = ["nynw", "ocer", "whfa", "youngst", "lvrpool", "memdfu", "hilloh", "grovoh", "mentoh", "encinogho",
                  "lahdfu"]
    s3_site = ["rsci"]
    total_site = local_site + s3_site
    site_list = {}
    for i in total_site:
        site_list[i] = {}
        if i in local_site:
            site_list[i]["type"] = "local"
        else:
            site_list[i]["type"] = "s3"
    sql_path = sql_get.dfu_sql_find(site_list)
    for i in total_site:
        site_list[i]["sql_path"] = sql_path[i]
        print(i + " " + sql_path[i])

    a = input("is the path right? ")
    if a !="no":
        return site_list
    else:
        return False


def output_total():
    site_list = refresh_sql_database()
    if site_list != False:
        check_site = ["nynw", "ocer", "whfa", "youngst", "lvrpool", "memdfu", "hilloh", "grovoh", "mentoh", "encinogho",
                      "lahdfu", "rsci"]
        check_list = {}
        for check in check_site:
            check_list[check] = site_list[check]

        data_sites = []
        for i in check_list.keys():
            path = check_list[i]["sql_path"]
            df_guid, df_image, site, check_path = server_table_output(path)
            data_sites.append(df_guid)

        union_df = pd.concat(data_sites)
        union_df.to_excel("/Users/ziweishi/Desktop/dfu_check.xlsx")


        return union_df
    else:
        print("Wrong Path!")


def make_summary():

    df_type = pd.read_excel("/Users/ziweishi/Desktop/wausi_subject.xlsx",sheet_name="total")
    df_guid = pd.read_excel("/Users/ziweishi/Desktop/wausi_subject.xlsx",sheet_name="image")
    df = pd.merge(df_type,df_guid,on="MedicalNumber",how="left")

    df_tra = df[df["type"]=="training"]
    df_tra = df_tra.copy()
    df_tra["site"] = df_tra["site"].apply(lambda x: str(x))

    tra_img_total = df_tra["MedicalNumber"].to_list()
    tra_sub_image_count = Counter(tra_img_total)
    tra_img_sum = pd.DataFrame(list(tra_sub_image_count.items()), columns=["SubjectID", "Image_Num"])
    # print(tra_img_sum)

    tra_sub = tra_sub_image_count.keys()
    tra_site_ab = [x[0:3] for x in tra_sub]
    tra_site_info = {}
    for i in tra_sub:
        sub_tra = tra_img_sum[tra_img_sum["SubjectID"] == i]
        num = sub_tra["Image_Num"].iloc[0]
        ii = i[0:3]
        if ii not in tra_site_info:
            tra_site_info[ii] = num
        else:
            tra_site_info[ii] += num

    tra_site_num = Counter(tra_site_ab)
    tra_site_sum = pd.DataFrame(list(tra_site_num.items()), columns=["Site", "Sub_Num"])
    tra_site_list = tra_site_sum["Site"].to_list()

    tra_sum = []
    tra_site_name = []

    for j in tra_site_list:
        tra_sub_info = df_tra[df_tra["site"] == j]
        j_name = tra_sub_info["site_name"].iloc[0]
        sum_num = tra_site_info[j]
        tra_sum.append(sum_num)
        tra_site_name.append(j_name)

    tra_site_sum["Img_Sum"] = tra_sum
    tra_site_sum["Site_Name"] = tra_site_name

    tra_total ={}
    tra_total["Site"] = "total"
    tra_total["Sub_Num"] = len(tra_sub)
    tra_total["Img_Sum"] = len(tra_img_total)
    tra_total["Site_Name"]=np.nan

    print(tra_total)


    df_tra_total = pd.DataFrame(tra_total,index=[0])
    df_tra_list = [tra_site_sum,df_tra_total]
    df_tra_final = pd.concat(df_tra_list)







    df_val = df[df["type"] == "validation"]
    df_val = df_val.copy()
    df_val["site"] = df_val["site"].apply(lambda x: str(x))

    val_img_total = df_val["MedicalNumber"].to_list()
    val_sub_image_count = Counter(val_img_total)
    val_img_sum = pd.DataFrame(list(val_sub_image_count.items()), columns=["SubjectID", "Image_Num"])


    val_sub = val_sub_image_count.keys()
    val_site_ab = [x[0:3] for x in val_sub]
    val_site_info = {}
    for i in val_sub:
        sub_val = val_img_sum[val_img_sum["SubjectID"] == i]
        num = sub_val["Image_Num"].iloc[0]
        ii = i[0:3]
        if ii not in val_site_info:
            val_site_info[ii] = num
        else:
            val_site_info[ii] += num

    val_site_num = Counter(val_site_ab)
    val_site_sum = pd.DataFrame(list(val_site_num.items()), columns=["Site", "Sub_Num"])
    val_site_list = val_site_sum["Site"].to_list()

    val_sum = []
    val_site_name = []

    for j in val_site_list:
        val_sub_info = df_val[df_val["site"] == j]
        j_name = val_sub_info["site_name"].iloc[0]
        sum_num = val_site_info[j]
        val_sum.append(sum_num)
        val_site_name.append(j_name)

    val_site_sum["Img_Sum"] = val_sum
    val_site_sum["Site_Name"] = val_site_name

    val_total = {}
    val_total["Site"] = "total"
    val_total["Sub_Num"] = len(val_sub)
    val_total["Img_Sum"] = len(val_img_total)
    val_total["Site_Name"] = np.nan

    df_val_total = pd.DataFrame(val_total,index=[0])
    df_val_list = [val_site_sum, df_val_total]
    df_val_final = pd.concat(df_val_list)

    writer = pd.ExcelWriter('/Users/ziweishi/Desktop/WAUSI_Summary.xlsx', engine='xlsxwriter')
    df_tra_final.to_excel(writer, sheet_name='training_site_summary', index=False)
    tra_img_sum.to_excel(writer, sheet_name='training_subject_summary', index=False)
    df_val_final.to_excel(writer, sheet_name='validation_site_summary', index=False)
    val_img_sum.to_excel(writer, sheet_name='validation_subject_summary', index=False)

    writer.close()


    workbook =load_workbook("/Users/ziweishi/Desktop/WAUSI_Summary.xlsx")

    # 对每个 sheet 设置单元格居中
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 保存更改后的 Excel 文件
    workbook.save("/Users/ziweishi/Desktop/WAUSI_Summary.xlsx")


if __name__ =="__main__":
    # make_summary()

    site_list = refresh_sql_database()
    # local_site = ["nynw", "ocer", "whfa", "youngst", "lvrpool", "memdfu", "hilloh", "grovoh", "mentoh", "encinogho","lahdfu","rsci"]


    if site_list !=False:
        check_site = ["nynw", "ocer", "whfa", "youngst", "lvrpool", "memdfu", "hilloh", "grovoh", "mentoh", "encinogho","lahdfu","rsci"]
        # check_site = ["memdfu"]
        check_list = {}
        for check in check_site:
            check_list[check] = site_list[check]

        db = download_whole_dynamodb_table.download_table("DFU_Master_ImageCollections")
        db = db[db["StudyName"] == "DFU_SSP"]

        data_sites = []
        for i in check_list.keys():
            path = check_list[i]["sql_path"]
            # print(path)
            df_guid, df_image, site, check_path = server_table_output(path)
            site_image = image_check(db, df_guid, df_image, site, check_path)
            data_sites.append(df_guid)

        union_df = pd.concat(data_sites)
        union_df.to_excel("/Users/ziweishi/Desktop/dfu_check.xlsx")

    else:
        print("Wrong Path!")












